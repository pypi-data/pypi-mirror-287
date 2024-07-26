import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class XlsxUtil:
    """
    A utility class for handling operations related to Excel files.
    """

    @staticmethod
    def saveExcel(datas,path,fileName):
        """
        Save data to an Excel file.

        Args:
            datas (list or dict): The data to be saved, which could be a list of dictionaries or a dictionary of lists.
            path (str): The directory path where the Excel file will be saved.
            fileName (str): The name of the Excel file (without the extension).

        Returns:
            None

        Example usage：
            XlsxUtil.saveExcel(data, '/path/to/save/', 'example_filename')
        """
        fullFileName = path+fileName+'.xlsx'
        default_sheet_name = 'sheet0'

        df = pd.DataFrame(datas)
        df.to_excel(fullFileName,sheet_name=default_sheet_name, index=False)

        wb = load_workbook(fullFileName)
        ws = wb.active
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # 获取列字母
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        wb.save(fullFileName)
