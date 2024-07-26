import json
import os

import openpyxl
import xlrd
from common.plugin.data_bus import DataBus
from datetime import datetime
from xlrd import xldate_as_tuple
from xlutils.copy import copy
from common.file.handle_system import adjust_path_data,adjust_path
from common.data.handle_common import get_system_key, format_caseName,print_info,print_debug
from common.plat.service_platform import ServicePlatForm
from common.common.constant import Constant
from loguru import logger
from common.data.data_process import DataProcess


def is_josn(inStr):
    try:
        json.loads(inStr)
    except:#只有try报错才执行里面的代码
        return False
    return True

def is_dict(inStr):
    try:
        type(inStr) == dict
    except:#只有try报错才执行里面的代码
        return False
    return True



def get_excel_data(excelPath,sheetIndex):
    """
    读取excel文件的方法
    :param excelPath: 文件路径
    :param sheetName: sheet页的名称
    :param caseName: 执行的case名称
    :return:
    """
    # 存放字典格式，{sheet1页的名称:[[第一行数据],[第二行数据][第三行数据]],sheet2页的名称:[[第一行数据],[第二行数据][第三行数据]]}
    sheet_dict = {}
    excelPath=adjust_path_data(excelPath)
    # 打开excel文件，并且保持原有样式
    workBook = xlrd.open_workbook(filename=excelPath,formatting_info=True)
    # Frame().logger(fileLog=True).info(f'获取用例路径为:{excelPath}的文件')
    # 拿到sheet页的内容
    table = workBook.sheet_by_index(sheetIndex)
    # 存放列表格式,，[[第一行数据],[第二行数据][第三行数据]],sheet2页的名称:[[第一行数据],[第二行数据][第三行数据]]
    sheet_datas = []
    # print(table)
    # table.nrows是总行数，按行号进行遍历，(从第一行开始，总行数)
    for row in range(1,table.nrows):
        # 定义一个空列表，存放的是每一行的数据
        row_data = []
        # 得到case状态
        cell_data = table.cell_value(row,4)
        # 如果case状态为无效，就跳过此次循环，不将此行数据加入row_data
        if cell_data =='否':
            continue
        # table.ncols是总列数，按列号进行遍历，(从第0列开始，总列)

        for col in range(0,table.ncols):
            # 获取指定单元格数据(行，列)
            cell_data = table.cell_value(row,col)
            # 如果是请求头、请求参数和响应参数这三列数据
            if col in (6,7):
                try:
                    cell_data = cell_data.replace('\n','').replace(' ','')
                    if is_josn(cell_data):  # 判断当前单元格数据是否是json
                        cell_data = json.loads(cell_data)  # 转化为字典格式
                        # print(f'cell_data:{cell_data}的类型是：{type(cell_data)}')
                except Exception as error:
                    raise error
            if col in (3,5):
                cell_data = cell_data.split('\n')

            # 将当前行所有单元格的数据追加进行数据列表
            row_data.append(cell_data)
            if col in (4,):
                row_data.remove(cell_data)
        # 将当前行列表数据追加到sheet页数据列表
        sheet_datas.append(row_data)
    return sheet_datas
    # Frame().logger(fileLog=True).info(f'获取指定sheet页：{sheetName}的所有行数据')


def writExcle(filePath,exlSheetName,dataLists,exlRow,exlColumn):
    """
    filePath:文件路径+文件名（后缀支持.xlsx），举例："C:/Users/qinyan5/Desktop/bagtestData.xlsx"
    exlSheetName：插入表格sheet页name
    dataLists：输入的数据data，格式：[[],[],[]]，举例：[['aaa','bbb','ccc'],['ggg','hhh','kkk'],['11','22','333']]
    需要的写入第几行，第几列
    exlRow:从表格的第几行开始输入，坐标：行
    exlColumn：从表格的第几列开始输入，坐标：列
    """
    filePath = adjust_path_data(filePath)
    dataListlen = len(dataLists)
    datalen = len(dataLists[0])
    wb = openpyxl.load_workbook(filePath)
    sh = wb[exlSheetName]
    for i in range(0,dataListlen):
        data = dataLists[i]
        for j in range(0,datalen):
            sh.cell(i+exlRow, j+exlColumn, data[j])
    wb.save(filePath)

#根据excle单元格列坐标批量读出数据
def readExcle(filePath,sheetName,columStart,columnX):
    """
    :param filePath: 文件路径
    :param sheetName: excle sheet页名称
    :param columStart: 开始取列码 :B2
    :param columnX: 列标记:B
    :return:
    """

    #打开excle文件
    filePath = adjust_path_data(filePath)
    wb = openpyxl.load_workbook(filePath)
    #sheet页
    sh = wb[sheetName]
    #获取最大列表数据行数
    max_row = (sh.max_row) -1
    # print(max_row)
    #获取需要去字段的单元格区间:sh[B2:B7]
    columnX = sh[columStart:columnX+'%d' % max_row]

    columnValueList = []
    #获取该列对应的所有的单元格对象
    for column_cells in columnX:
        for cell in column_cells:
            columnvalue = cell.value
            columnValueList.append(columnvalue)
    # print(columnValueList)
    return columnValueList


def readAllList(filePath,sheetName):
    """
    取全表数据，第一行作为title，每一行作为value
    :param filePath:
    :param sheetName:
    :return:
    """
    filePath = adjust_path_data(filePath)
    wb = openpyxl.load_workbook(filePath)
    sh = wb[sheetName]
    #表格第一行内容为title
    title = [i.value for i in sh[1]]
    max_row = sh.max_row + 1

    #key-->value
    data_list = []
    for j in range(2, max_row):
        data = [''if i.value == None else i.value for i in sh[j] ]
        datadict = dict(zip(title,data))
        data_list.append(datadict)
    return data_list

def readValueList(filePath,sheetName):
    """
    取全部sheet页数据，除第一行，其他作为list value
    :param filePath:
    :param sheetName:
    :return:
    """
    filePath = adjust_path_data(filePath)
    wb = openpyxl.load_workbook(filePath)
    sh = wb[sheetName]

    max_row = sh.max_row

    dataAll_list = []
    for i in range(2,max_row):
        data = ['' if j.value == None else j.value for j in sh[i]]

        dataAll_list.append(data)
    return dataAll_list




def readExcleMerge(filePath,sheetName):
    """
    取全表合并单元格坐标
    :param filePath:
    :param sheetName:
    :return:
    """
    filePath = adjust_path_data(filePath)
    wb = openpyxl.load_workbook(filePath)
    sh = wb[sheetName]
    exlSpace = sh.merged_cells.ranges

    for spaceData in exlSpace:
        datavalue = spaceData.start_cell.value
        # print(datavalue)
        return datavalue

def excel_to_list(data_file, sheet, _index:int=1, _filter:dict=None,  _replace: bool=True):
    """
    读取Excel中特定sheet的数据，按行将数据存入数组datalist
    :param data_file:Excel文件目录
    :param sheet:需要读取的sheet名称
    :return:datalist
    """
    data_file = adjust_path_data(data_file)
    data_list = []
    wb = xlrd.open_workbook(data_file)
    sh = wb.sheet_by_name(sheet)
    for row in range(_index, sh.nrows):
        d = dict()
        for col in range(0, sh.ncols):
            cell_data = sh.cell_value(row, col)
            ctype = sh.cell(row, col).ctype
            if ctype == 2 and cell_data % 1 == 0.0:
                cell_data = int(cell_data)
            if ctype == 3:
                date = datetime(*xldate_as_tuple(cell_data, 0))
                cell_data = date.strftime('%Y-%m-%d')
            col_title = sh.cell_value(0, col).strip()
            d[col_title] = cell_data
            d['$'+col_title] = cell_data
            d[col_title+'_cell'] = {"row": row, "col": col}
        if _replace:
            temp = DataBus.get_data(d)
            temp['_excelPath'] = data_file
            temp['_sheetName'] = sheet
        else:
            temp = d
            temp['_excelPath'] = data_file
            temp['_sheetName'] = sheet
        print_info(f'添加单个参数化用例数据: {temp}')
        if DataProcess.isNotNull(_filter):
            for k, v in _filter.items():
                if temp[k] == v:
                    data_list.append(temp)
        else:
            data_list.append(temp)
        print_info(f'用例参数化数据: {data_list}')
    return data_list

def check_excel_data(testdata:dict, _filter, _checkData):
    if _checkData:
        casename = format_caseName(testdata[Constant.CASE_TITLE])
        caseNo = testdata[Constant.CASE_NO]
        if DataProcess.isNotNull(get_system_key(Constant.TEST_SRTCYCLE_ID)):
            cycleId = get_system_key(Constant.TEST_SRTCYCLE_ID)
            _list = ServicePlatForm.getCaseRun(cycleId, caseNo, casename)
            if _list['status'] == Constant.STATUS_PRE or _list['status'] == Constant.STATUS_AUTOTEST:
                ServicePlatForm.updateByRunid(_list['caserunid'], Constant.STATUS_AUTOTEST_PARA, "")
                print_debug(f'用例信息:{str(_list)} 添加到执行队列')
                return True
            else:
                print_info(f'用例信息:{str(_list)} 添加到执行队列【无需添加】')
                return False
        if testdata[Constant.CASE_STATUS].strip() == '否':
            return False
        if DataProcess.isNotNull(get_system_key(Constant.TEST_CASE_NAME_LIST)):
            if testdata[Constant.CASE_TITLE] in eval(get_system_key(Constant.TEST_CASE_NAME_LIST)):
                return True
            else:
                return False
    if DataProcess.isNotNull(_filter):
        for k, v in _filter.items():
            if testdata[k] == v:
                return True
        return False
    return True


def write_to_excel(excel_name, sheet_name, row, col, content):
    """
    写数据进excel
    :param excel_name: 文件名  data.xls
    :param sheet_name: sheet名称
    :param row: 行
    :param col: 列
    :param content: 内容
    :return:
    """
    file_path = adjust_path(excel_name)
    # 打开jira的xls文件
    jira_rb = xlrd.open_workbook(file_path, formatting_info=True)
    jira_wb = copy(jira_rb)
    # 获取指定sheet页对象
    jira_sheet = jira_wb.get_sheet(sheet_name)
    # 写入内容
    jira_sheet.write(row, col, content)
    # 保存
    jira_wb.save(file_path)


def get_test_data(data_list, case_name):
    """
    根据用例名称查找用例本身
    :param data_list:按行存储用例的数组
    :param case_name: 用例名称
    :return: case_data
    """
    for case_data in data_list:
        # 如果字典数据中case_name与参数一致
        if case_name == case_data['用例标题']:
            return case_data
        # 如果查询不到会返回None
        else:
            return None


