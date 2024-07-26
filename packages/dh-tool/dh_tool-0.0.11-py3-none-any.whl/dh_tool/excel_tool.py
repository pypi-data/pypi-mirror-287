import pandas as pd
from openpyxl.utils import get_column_letter
import openpyxl


class ExcelTool:
    def __init__(self) -> None:
        self.width_config = {
            "Comments": {"width": 90},
            "BestSentence1": {"width": 20},
            "BestSentence2": {"width": 20},
            "FeedBack": {"width": 40},
            "timestamp": {"width": 20},
            "level": {"width": 10},
            "topic": {"width": 20},
            "message": {"width": 40},
            "description": {"width": 60},
            "traceback": {"width": 80},
        }

    def setup_width_config(self, **kwargs):
        self.width_config.update(kwargs)
        print(f"Config updated to {self.width_config}")

    def post_excel(self, df: pd.DataFrame, save_path: str, default_width=6):
        if save_path.endswith(".xlsx"):
            save_path = save_path.rstrip(".xlsx")
        writer = pd.ExcelWriter(f"{save_path}.xlsx", engine="openpyxl")
        df.to_excel(writer, index=False, sheet_name="Sheet1")

        # 엑셀 시트 가져오기
        worksheet = writer.sheets["Sheet1"]

        # 열 너비 자동 조정 또는 사용자 지정
        for column in df.columns:
            # column_width = max(df[column].astype(str).apply(len).max(), len(column))
            column_width = default_width
            if column in self.width_config:
                if "width" in self.width_config[column]:
                    column_width = self.width_config[column]["width"]
            worksheet.column_dimensions[
                get_column_letter(df.columns.get_loc(column) + 1)
            ].width = column_width

        # 첫 번째 행 고정
        worksheet.freeze_panes = "A2"

        # 자동 줄 바꿈 적용
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=worksheet.max_column,
        ):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

        # 변경 사항 저장하고 파일 닫기
        writer.close()
