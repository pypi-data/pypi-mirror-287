"""Support for reading tabular data"""

import csv
from io import BytesIO, StringIO
import os

from fastapi import UploadFile
from openpyxl import load_workbook

from ..sys_utils.log_tools import setup_logger

logger = setup_logger(__name__)


def read_data_table(file: str | UploadFile) -> list[dict[str, str]]:
    """
    Reads a file (CSV or XLSX) and returns a list of per-row dicts.

    Args:
        file (str | UploadFile): The file path or UploadFile object.

    Returns:
        list[dict[str, str]]: List of dictionaries where each dict represents a row.
    """
    if hasattr(file, "filename"):
        file_extension = os.path.splitext(file.filename)[1]
    elif isinstance(file, str):
        _, file_extension = os.path.splitext(file)
    else:
        logger.error(f"Can't handle file object {file}")
        raise ValueError("Unsupported file object")

    if file_extension.lower() == ".xlsx":
        return read_xlsx_data_table(file.file)
    if file_extension.lower() == ".csv":
        return read_csv_data_table(file.file)
    raise ValueError("Unsupported file type")


def read_xlsx_data_table(file: str | BytesIO) -> list[dict[str, str]]:
    """
    Reads an XLSX file and returns a list of per-row dicts.

    Args:
        file (str | BytesIO): The file path or file-like object.

    Returns:
        list[dict[str, str]]: List of dictionaries where each dict represents a row.
    """
    if isinstance(file, str):
        workbook = load_workbook(filename=file, data_only=True)
    else:
        workbook = load_workbook(file, data_only=True)

    sheet = workbook.active
    headers = transform_headers([cell.value for cell in sheet[1]])

    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = {
            headers[i]: str(value) if value is not None else ""
            for i, value in enumerate(row)
        }
        rows.append(row_dict)

    return rows


def read_csv_data_table(file: str | StringIO | BytesIO) -> list[dict[str, str]]:
    """
    Reads a CSV file and returns a list of per-row dicts.

    Args:
        file (str | StringIO | BytesIO): The file path or file-like object.

    Returns:
        list[dict[str, str]]: List of dictionaries where each dict represents a row.
    """
    if isinstance(file, str):
        with open(file, mode="r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            reader.fieldnames = transform_headers(reader.fieldnames)
            return [row for row in reader]
    else:
        content = file.read().decode("utf-8").splitlines()
        reader = csv.DictReader(content)
        reader.fieldnames = transform_headers(reader.fieldnames)
        return [row for row in reader]


def transform_headers(headers: list[str]) -> list[str]:
    """
    Transforms the headers of the input file to a consistent format.
    """
    return [
        {
            "Case Code": "case_code",
            " document Title": "doc_title",
            " document Comment": "doc_comment",
        }.get(header, header)
        for header in headers
    ]
