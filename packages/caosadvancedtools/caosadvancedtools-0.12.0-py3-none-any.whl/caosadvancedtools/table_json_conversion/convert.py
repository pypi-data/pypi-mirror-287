# encoding: utf-8
#
# This file is a part of the LinkAhead Project.
#
# Copyright (C) 2024 Indiscale GmbH <info@indiscale.com>
# Copyright (C) 2024 Daniel Hornung <d.hornung@indiscale.com>
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.

"""Convert XLSX files to JSON dictionaries."""

from __future__ import annotations

import datetime
import itertools
import sys
from functools import reduce
from operator import getitem
from types import SimpleNamespace
from typing import Any, BinaryIO, Callable, TextIO, Union
from warnings import warn

import jsonschema
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from caosadvancedtools.table_json_conversion import xlsx_utils
from caosadvancedtools.table_json_conversion.fill_xlsx import read_or_dict


def _strict_bool(value: Any) -> bool:
    """Convert value to bool, but only if it really is a valid XLSX bool."""
    if isinstance(value, bool):
        return value
    raise TypeError(f"Not a good boolean: {repr(value)}")


class ForeignError(KeyError):
    def __init__(self, *args, definitions: list, message: str = ""):
        super().__init__(message, *args)
        self.definitions = definitions


class XLSXConverter:
    """Class for conversion from XLSX to JSON.

For a detailed description of the required formatting of the XLSX files, see ``specs.md`` in the
documentation.
    """

    PARSER: dict[str, Callable] = {
        "string": str,
        "number": float,
        "integer": int,
        "boolean": _strict_bool,
    }

    def __init__(self, xlsx: Union[str, BinaryIO], schema: Union[dict, str, TextIO],
                 strict: bool = False):
        """
Parameters
----------
xlsx: Union[str, BinaryIO]
  Path to the XLSX file or opened file object.

schema: Union[dict, str, TextIO]
  Schema for validation of XLSX content.

strict: bool, optional
  If True, fail faster.
"""
        self._workbook = load_workbook(xlsx)
        self._schema = read_or_dict(schema)
        self._defining_path_index = xlsx_utils.get_defining_paths(self._workbook)
        self._check_columns(fail_fast=strict)
        self._handled_sheets: set[str] = set()
        self._result: dict = {}
        self._errors: dict = {}

    def to_dict(self, validate: bool = False, collect_errors: bool = True) -> dict:
        """Convert the xlsx contents to a dict.

Parameters
----------
validate: bool, optional
  If True, validate the result against the schema.

collect_errors: bool, optional
  If True, do not fail at the first error, but try to collect as many errors as possible.  After an
  Exception is raised, the errors can be collected with ``get_errors()`` and printed with
  ``get_error_str()``.

Returns
-------
out: dict
  A dict representing the JSON with the extracted data.
        """
        self._handled_sheets = set()
        self._result = {}
        self._errors = {}
        for sheetname in self._workbook.sheetnames:
            if sheetname not in self._handled_sheets:
                self._handle_sheet(self._workbook[sheetname], fail_later=collect_errors)
        if validate:
            jsonschema.validate(self._result, self._schema)
        if self._errors:
            raise RuntimeError("There were error while handling the XLSX file.")
        return self._result

    def get_errors(self) -> dict:
        """Return a dict with collected errors."""
        return self._errors

    def get_error_str(self) -> str:
        """Return a beautiful string with the collected errors."""
        result = ""
        for loc, value in self._errors.items():
            result += f"Sheet: {loc[0]}\tRow: {loc[1] + 1}\n"
            for item in value:
                result += f"\t\t{item[:-1]}:\t{item[-1]}\n"
        return result

    def _check_columns(self, fail_fast: bool = False):
        """Check if the columns correspond to the schema."""
        def missing(path):
            message = f"Missing column: {xlsx_utils.p2s(path)}"
            if fail_fast:
                raise ValueError(message)
            else:
                warn(message)
        for sheetname in self._workbook.sheetnames:
            sheet = self._workbook[sheetname]
            parents: dict = {}
            col_paths = []
            for col in xlsx_utils.get_data_columns(sheet).values():
                parents[xlsx_utils.p2s(col.path[:-1])] = col.path[:-1]
                col_paths.append(col.path)
            for path in parents.values():
                subschema = xlsx_utils.get_subschema(path, self._schema)

                # Unfortunately, there are a lot of special cases to handle here.
                if subschema.get("type") == "array":
                    subschema = subschema["items"]
                if "enum" in subschema:  # Was handled in parent level already
                    continue
                for child, content in subschema["properties"].items():
                    child_path = path + [child]
                    if content == {'type': 'string', 'format': 'data-url'}:
                        continue  # skip files
                    if content.get("type") == "array" and (
                            content.get("items").get("type") == "object"):
                        if child_path not in itertools.chain(*self._defining_path_index.values()):
                            missing(child_path)
                    elif content.get("type") == "array" and "enum" in content.get("items", []) and (
                            content.get("uniqueItems") is True):
                        # multiple choice
                        for choice in content["items"]["enum"]:
                            if child_path + [choice] not in col_paths:
                                missing(child_path + [choice])
                    elif content.get("type") == "object":
                        pass
                    else:
                        if child_path not in col_paths:
                            missing(child_path)

    def _handle_sheet(self, sheet: Worksheet, fail_later: bool = False) -> None:
        """Add the contents of the sheet to the result (stored in ``self._result``).

Each row in the sheet corresponds to one entry in an array in the result.  Which array exactly is
defined by the sheet's "proper name" and the content of the foreign columns.

Look at ``xlsx_utils.get_path_position`` for the specification of the "proper name".


Parameters
----------
fail_later: bool, optional
  If True, do not fail with unresolvable foreign definitions, but collect all errors.
"""
        row_type_column = xlsx_utils.get_row_type_column_index(sheet)
        foreign_columns = xlsx_utils.get_foreign_key_columns(sheet)
        foreign_column_paths = {col.index: col.path for col in foreign_columns.values()}
        data_columns = xlsx_utils.get_data_columns(sheet)
        data_column_paths = {col.index: col.path for col in data_columns.values()}
        # Parent path, insert in correct order.
        parent, proper_name = xlsx_utils.get_path_position(sheet)
        if parent:
            parent_sheetname = xlsx_utils.get_worksheet_for_path(parent, self._defining_path_index)
            if parent_sheetname not in self._handled_sheets:
                self._handle_sheet(self._workbook[parent_sheetname], fail_later=fail_later)

        # # We save single entries in lists, indexed by their foreign key contents.  Each entry
        # # consists of:
        # # - foreign: Dict with path -> value for the foreign columns
        # # - data: The actual data of this entry, a dict.
        # entries: dict[str, list[SimpleNamespace]] = {}

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            # Skip non-data rows.
            if row[row_type_column] is not None:
                continue
            foreign_repr = ""
            foreign = []  # A list of lists, each of which is: [path1, path2, ..., leaf, value]
            data: dict = {}     # Local data dict
            # Collect data (in dict relative to current level) and foreign data information
            for col_idx, value in enumerate(row):
                if col_idx in foreign_column_paths:
                    foreign_repr += str(value)
                    foreign.append(foreign_column_paths[col_idx] + [value])
                    continue

                if col_idx in data_column_paths:
                    path = data_column_paths[col_idx]
                    if self._is_multiple_choice(path):
                        real_value = path.pop()  # Last component is the enum value, insert above
                        # set up list
                        try:
                            _set_in_nested(mydict=data, path=path, value=[], prefix=parent, skip=1)
                        except ValueError as err:
                            if not str(err).startswith("There is already some value at"):
                                raise
                        if not xlsx_utils.parse_multiple_choice(value):
                            continue
                        _set_in_nested(mydict=data, path=path, value=real_value, prefix=parent,
                                       skip=1, append_to_list=True)
                    else:
                        value = self._validate_and_convert(value, path)
                        _set_in_nested(mydict=data, path=path, value=value, prefix=parent, skip=1)
                    continue

            try:
                # Find current position in tree
                parent_dict = self._get_parent_dict(parent_path=parent, foreign=foreign)

                # Append data to current position's list
                if proper_name not in parent_dict:
                    parent_dict[proper_name] = []
                parent_dict[proper_name].append(data)
            except ForeignError as kerr:
                if not fail_later:
                    raise
                self._errors[(sheet.title, row_idx)] = kerr.definitions
        self._handled_sheets.add(sheet.title)

    def _is_multiple_choice(self, path: list[str]) -> bool:
        """Test if the path belongs to a multiple choice section."""
        if not path:
            return False
        subschema = self._get_subschema(path[:-1])
        if (subschema["type"] == "array"
                and subschema.get("uniqueItems") is True
                and "enum" in subschema["items"]):
            return True
        return False

    def _get_parent_dict(self, parent_path: list[str], foreign: list[list]) -> dict:
        """Return the dict into which values can be inserted.

This method returns, from the current result-in-making, the entry at ``parent_path`` which matches
the values given in the ``foreign`` specification.
"""
        foreign_groups = _group_foreign_paths(foreign, common=parent_path)

        current_object = self._result
        for group in foreign_groups:
            # Find list for which foreign definitions are relevant.
            current_object = reduce(getitem, group.subpath, current_object)
            assert isinstance(current_object, list)
            # Test all candidates.
            for cand in current_object:
                if all(reduce(getitem, definition[:-1], cand) == definition[-1]
                       for definition in group.definitions):
                    current_object = cand
                    break
            else:
                message = f"Cannot find an element at {parent_path} for these foreign defs:\n"
                for name, value in group.definitions:
                    message += f"    {name}: {value}\n"
                print(message, file=sys.stderr)
                error = ForeignError(definitions=group.definitions, message=message)
                raise error
        assert isinstance(current_object, dict)
        return current_object

    def _validate_and_convert(self, value: Any, path: list[str]):
        """Apply some basic validation and conversion steps.

This includes:
- Validation against the type given in the schema
- List typed values are split at semicolons and validated individually
        """
        if value is None:
            return value
        subschema = self._get_subschema(path)
        # Array handling only if schema says it's an array.
        if subschema.get("type") == "array":
            array_type = subschema["items"]["type"]
            if isinstance(value, str) and ";" in value:
                values = [self.PARSER[array_type](v) for v in value.split(";")]
                return values
        try:
            # special case: datetime or date
            if ("anyOf" in subschema):
                if isinstance(value, datetime.datetime) and (
                        {'type': 'string', 'format': 'date-time'} in subschema["anyOf"]):
                    return value
                if isinstance(value, datetime.date) and (
                        {'type': 'string', 'format': 'date'} in subschema["anyOf"]):
                    return value
            jsonschema.validate(value, subschema)
        except jsonschema.ValidationError as verr:
            print(verr)
            print(path)
            raise

        # Finally: convert to target type
        return self.PARSER[subschema.get("type", "string")](value)

    def _get_subschema(self, path: list[str], schema: dict = None) -> dict:
        """Return the sub schema at ``path``."""
        if schema is None:
            schema = self._schema
            assert schema is not None
        assert isinstance(schema, dict)

        return xlsx_utils.get_subschema(path, schema)


def _group_foreign_paths(foreign: list[list], common: list[str]) -> list[SimpleNamespace]:
    """Group the foreign keys by their base paths.

Parameters
----------
foreign: list[list]
  A list of foreign definitions, consisting of path components, property and possibly value.

common: list[list[str]]
  A common path which defines the final target of the foreign definitions.  This helps to understand
  where the ``foreign`` paths shall be split.

Returns
-------
out: list[dict[str, list[list]]]

  A list of foreign path segments, grouped by their common segments.  Each element is a namespace
  with detailed information of all those elements which form the group.  The namespace has the
  following attributes:

  - ``path``: The full path to this path segment.  This is always the previous segment's ``path``
    plus this segment's ``subpath``.
  - ``stringpath``: The stringified ``path``, might be useful for comparison or sorting.
  - ``subpath``: The path, relative from the previous segment.
  - ``definitions``: A list of the foreign definitions for this segment, but stripped of the
    ``path`` components.
    """
    # Build a simple dict first, without subpath.
    results = {}
    for f_path in foreign:
        path = []
        for component in f_path:
            path.append(component)
            if path != common[:len(path)]:
                break
        path.pop()
        definition = f_path[len(path):]
        stringpath = xlsx_utils.p2s(path)
        if stringpath not in results:
            results[stringpath] = SimpleNamespace(stringpath=stringpath, path=path,
                                                  definitions=[definition])
        else:
            results[stringpath].definitions.append(definition)

    # Then sort by stringpath and calculate subpath.
    stringpaths = sorted(results.keys())

    resultlist = []
    last_level = 0
    for stringpath in stringpaths:
        elem = results[stringpath]
        elem.subpath = elem.path[last_level:]
        last_level = len(elem.path)
        resultlist.append(elem)

    # from IPython import embed
    # embed()

    if last_level != len(common):
        raise ValueError("Foreign keys must cover the complete `common` depth.")
    return resultlist


# pylint: disable-next=dangerous-default-value,too-many-arguments
def _set_in_nested(mydict: dict, path: list, value: Any, prefix: list = [], skip: int = 0,
                   overwrite: bool = False, append_to_list: bool = False) -> dict:
    """Set a value in a nested dict.

Parameters
----------
mydict: dict
  The dict into which the ``value`` shall be inserted.
path: list
  A list of keys, denoting the location of the value.
value
  The value which shall be set inside the dict.
prefix: list
  A list of keys which shall be removed from ``path``.  A KeyError is raised if ``path`` does not
  start with the elements of ``prefix``.
skip: int = 0
  Remove this many additional levels from the path, *after* removing the prefix.
overwrite: bool = False
  If True, allow overwriting existing content. Otherwise, attempting to overwrite existing values
  leads to an exception.
append_to_list: bool = False
  If True, assume that the element at ``path`` is a list and append the value to it.  If the list
  does not exist, create it.  If there is a non-list at ``path`` already, overwrite it with a new
  list, if ``overwrite`` is True, otherwise raise a ValueError.

Returns
-------
mydict: dict
  The same dictionary that was given as a parameter, but modified.
    """
    for idx, el in enumerate(prefix):
        if path[idx] != el:
            raise KeyError(f"Path does not start with prefix: {prefix} not in {path}")
    path = path[len(prefix):]
    if skip:
        assert len(path) > skip, f"Path must be long enoug to remove skip={skip} elements."
        path = path[skip:]

    tmp_dict = mydict
    while len(path) > 1:
        key = path.pop(0)
        if key not in tmp_dict:
            tmp_dict[key] = {}
        if not isinstance(tmp_dict[key], dict):
            if overwrite:
                tmp_dict[key] = {}
            else:
                raise ValueError(f"There is already some value at {path}")
        tmp_dict = tmp_dict[key]
    key = path.pop()
    if append_to_list:
        if key not in tmp_dict:
            tmp_dict[key] = []
        if not isinstance(tmp_dict[key], list):
            if overwrite:
                tmp_dict[key] = []
            else:
                raise ValueError(f"There is already some non-list value at [{key}]")
        tmp_dict[key].append(value)
    else:
        if key in tmp_dict and not overwrite:
            raise ValueError(f"There is already some value at [{key}]")
        if key not in tmp_dict:
            tmp_dict[key] = {}
        tmp_dict[key] = value
    return mydict


def to_dict(xlsx: Union[str, BinaryIO], schema: Union[dict, str, TextIO],
            validate: bool = None, strict: bool = False) -> dict:
    """Convert the xlsx contents to a dict, it must follow a schema.

Parameters
----------
xlsx: Union[str, BinaryIO]
  Path to the XLSX file or opened file object.

schema: Union[dict, str, TextIO]
  Schema for validation of XLSX content.

validate: bool, optional
  If True, validate the result against the schema.

strict: bool, optional
  If True, fail faster.


Returns
-------
out: dict
  A dict representing the JSON with the extracted data.
    """
    converter = XLSXConverter(xlsx, schema, strict=strict)
    return converter.to_dict()
